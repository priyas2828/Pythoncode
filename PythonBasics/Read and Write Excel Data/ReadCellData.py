
import openpyxl

# Load workbook
wk = openpyxl.load_workbook("C:\\developement\\ReadandWriteData.xlsx")

# First appraoch is to read data from excel
# Pass the sheet name
sh = wk['Sheet1'] # Whatever the sheet name you are giving this will return as list so it should be [] brackets
# Read the data from A3 from sheet
# print(sh['A3'].value)
# print(sh['B4'].value)

# 2. Creating object of cell

c1 = sh.cell(3, 1) # row and column
print(c1.value)

# 3. Keyword argument
c2 = sh.cell(column=1, row=3) # Keyword argument and result is still same
print(c2.value)

# Now to fetch the  C1 object i.e which column and which row

print(c1.row)
print(c1.column)