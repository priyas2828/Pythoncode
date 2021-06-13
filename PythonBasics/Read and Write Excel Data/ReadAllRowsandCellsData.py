import openpyxl
# Load workbook
wk = openpyxl.load_workbook("C:\\developement\\ReadandWriteData.xlsx")
sh = wk['Sheet1']

# Find rows having data
rows = sh.max_row
columns = sh.max_column

print("Total Rows are : " + str(rows))
print("Total Columns are : " + str(columns))
# First Appraoch:Now I want to read data from all the rows and columns
#for i in range(1, rows+1):
#    for j in range(1, columns+1):
#        c=sh.cell(i, j)
#        print(c.value)

# Second approach is
for r in sh['A1': 'C3']:  # In row we need to give the columns and rows you want
    for c in r:  # In clomuns , how many rows are there and fetch it and print c value

        print(c.value)

